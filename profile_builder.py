# SIMPLE PROFILE BUILDER
# Beginner Student Version

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ---------------- EXCEL FILE ----------------

file_name = "profile.xlsx"

# create excel file if not existing
if os.path.exists(file_name):
    workbook = load_workbook(file_name)

else:
    workbook = Workbook()
    sheet = workbook.active

    # headers
    sheet["A1"] = "ID"
    sheet["B1"] = "Last Name"
    sheet["C1"] = "First Name"
    sheet["D1"] = "Middle Name"
    sheet["E1"] = "Birth Year"
    sheet["F1"] = "Age"

    workbook.save(file_name)

# open excel
workbook = load_workbook(file_name)
sheet = workbook.active

# ---------------- FUNCTIONS ----------------

def display():

    # clear table
    for data in table.get_children():
        table.delete(data)

    # display excel data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)


def submit():

    fname = first_name_entry.get()
    mname = middle_name_entry.get()
    lname = last_name_entry.get()
    birth = birth_year_entry.get()

    # validation
    if fname == "" or lname == "" or birth == "":
        messagebox.showerror("Error", "Complete the form")
        return

    # age
    current_year = datetime.now().year
    age = current_year - int(birth)

    # id
    id_num = sheet.max_row

    # save to excel
    sheet.append([id_num, lname, fname, mname, birth, age])

    workbook.save(file_name)

    messagebox.showinfo("Saved", "Data Saved")

    # clear entry
    first_name_entry.delete(0, tk.END)
    middle_name_entry.delete(0, tk.END)
    last_name_entry.delete(0, tk.END)
    birth_year_entry.delete(0, tk.END)

    display()


def delete():

    selected = table.focus()

    if selected == "":
        return

    data = table.item(selected)
    values = data["values"]

    row_id = values[0]

    sheet.delete_rows(row_id + 1)

    workbook.save(file_name)

    display()

# ---------------- WINDOW ----------------

window = tk.Tk()
window.title("Age Calculator")
window.geometry("1000x500")
window.configure(bg="lightgreen")

# ---------------- TITLE ----------------

title = tk.Label(
    window,
    text="Profile Builder",
    font=("Times New Roman", 18, "bold"),
    bg="lightgreen"
)

title.pack(pady=10)

# ---------------- FORM ----------------

form = tk.Frame(window, bg="lightgreen")
form.pack()

# first name
first_name_entry = tk.Entry(form, width=25)
first_name_entry.grid(row=0, column=0, padx=5)

first_name_label = tk.Label(
    form,
    text="First Name",
    bg="lightgreen"
)

first_name_label.grid(row=1, column=0)

# middle name
middle_name_entry = tk.Entry(form, width=25)
middle_name_entry.grid(row=0, column=1, padx=5)

middle_name_label = tk.Label(
    form,
    text="Middle Name",
    bg="lightgreen"
)

middle_name_label.grid(row=1, column=1)

# last name
last_name_entry = tk.Entry(form, width=25)
last_name_entry.grid(row=0, column=2, padx=5)

last_name_label = tk.Label(
    form,
    text="Last Name",
    bg="lightgreen"
)

last_name_label.grid(row=1, column=2)

# birth year
birth_year_entry = tk.Entry(form, width=25)
birth_year_entry.grid(row=2, column=0, pady=10)

birth_label = tk.Label(
    form,
    text="Birth Year",
    bg="lightgreen"
)

birth_label.grid(row=3, column=0)

# ---------------- BUTTONS ----------------

button_frame = tk.Frame(window, bg="lightgreen")
button_frame.pack(pady=10)

submit_button = tk.Button(
    button_frame,
    text="Submit",
    bg="pink",
    command=submit
)

submit_button.grid(row=0, column=0, padx=20)

update_button = tk.Button(
    button_frame,
    text="Update",
    bg="lightgray"
)

update_button.grid(row=0, column=1, padx=20)

delete_button = tk.Button(
    button_frame,
    text="Delete",
    bg="red",
    fg="white",
    command=delete
)

delete_button.grid(row=0, column=2, padx=20)

# ---------------- TABLE ----------------

columns = (
    "ID",
    "Last Name",
    "First Name",
    "Middle Name",
    "Birth Year",
    "Age"
)

table = ttk.Treeview(
    window,
    columns=columns,
    show="headings",
    height=10
)

for col in columns:
    table.heading(col, text=col)
    table.column(col, width=140)

table.pack(fill="both", expand=True)

# display data
display()

window.mainloop()
