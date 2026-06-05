
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op
import os



# --- Functions ---

def display():
    workbook = op.load_workbook("medicine_inventory7.xlsx")
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)

def input_validation():
    name = name_entry.get()
    category = category_entry.get()
    quantity = quantity_entry.get()
    price = price_entry.get()
    expiry = expiry_entry.get()

    if not name or not category or not quantity or not price or not expiry:
        messagebox.showerror("Error","All fields required!")
        return False

    if not quantity.isdigit():
        messagebox.showerror("Error","Quantity must be a number")
        return False

    try:
        float(price)

    except:
        messagebox.showerror("Error","Price must be a number")
        return False
    return True

def saving():
    if not input_validation():
        return

    name = name_entry.get()
    category = category_entry.get()
    quantity = int(quantity_entry.get())
    price = float(price_entry.get())
    expiry = expiry_entry.get()

    workbook = op.load_workbook("medicine_inventory7.xlsx")
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([new_id,name,category,quantity,price,expiry])
    workbook.save("medicine_inventory7.xlsx")

    messagebox.showinfo("Success","Medicine added successfully!")

    clear_entries()
    display()

def auto_populate(event):
    selected = table.focus()
    values = table.item(selected,"values")

    if values:
        clear_entries()

        name_entry.insert(0, values[1])
        category_entry.insert(0, values[2])
        quantity_entry.insert(0, values[3])
        price_entry.insert(0, values[4])
        expiry_entry.insert(0, values[5])

def update():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error","Select a record first.")
        return

    if not input_validation():
        return

    values = table.item(selected,"values")
    record_id = values[0]
    workbook = op.load_workbook("medicine_inventory7.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == str(record_id):
            row[1].value = name_entry.get()
            row[2].value = category_entry.get()
            row[3].value = int(quantity_entry.get())
            row[4].value = float(price_entry.get())
            row[5].value = expiry_entry.get()
            break
    workbook.save("medicine_inventory7.xlsx")
    messagebox.showinfo("Success","Record updated!")

    display()

def delete_record():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error","Select a record.")
        return

    if messagebox.askyesno("Confirm","Delete this record?"):

        values = table.item(selected,"values")
        record_id = values[0]

        workbook = op.load_workbook("medicine_inventory7.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == str(record_id):
                sheet.delete_rows(row[0].row,1)
                break

        workbook.save("medicine_inventory7.xlsx")
        display()
        clear_entries()

        messagebox.showinfo("Success","Record deleted!")

def clear_entries():
    name_entry.delete(0,tk.END)
    category_entry.delete(0,tk.END)
    quantity_entry.delete(0,tk.END)
    price_entry.delete(0,tk.END)
    expiry_entry.delete(0,tk.END)

# --- GUI ---

window = tk.Tk()

window.title("Medicine Inventory System")
window.geometry("750x500")
title = tk.Label(window,text="MEDICINE INVENTORY SYSTEM",font=("Arial",16,"bold"))
title.pack()

frame = tk.Frame(window)
frame.pack()

tk.Label(frame,text="Medicine Name").grid(row=0,column=0)
name_entry = tk.Entry(frame)
name_entry.grid(row=0,column=1)

tk.Label(frame,text="Category").grid(row=1,column=0)
category_entry = tk.Entry(frame)
category_entry.grid(row=1,column=1)

tk.Label(frame,text="Quantity").grid(row=2,column=0)
quantity_entry = tk.Entry(frame)
quantity_entry.grid(row=2,column=1)

tk.Label(frame,text="Price").grid(row=3,column=0)
price_entry = tk.Entry(frame)
price_entry.grid(row=3,column=1)

tk.Label(frame,text="Expiry Date").grid(row=4,column=0)
expiry_entry = tk.Entry(frame)
expiry_entry.grid(row=4,column=1)


btn = tk.Frame(window)
btn.pack(pady=10)
tk.Button(btn,text="Add",width=10,command=saving).grid(row=0,column=0)
tk.Button(btn,text="Update",width=10,command=update).grid(row=0,column=1)
tk.Button(btn,text="Delete",width=10,command=delete_record).grid(row=0,column=2)
tk.Button(btn,text="Clear",width=10,command=clear_entries).grid(row=0,column=3)

columns = ("ID","Medicine","Category","Qty","Price","Exp:")
table = ttk.Treeview(window,columns=columns,show="headings")

for col in columns:
    table.heading(col,text=col)
    table.column(col,width=110)

table.pack(fill="both",expand=True)
table.bind("<ButtonRelease-1>",auto_populate)

window.mainloop()