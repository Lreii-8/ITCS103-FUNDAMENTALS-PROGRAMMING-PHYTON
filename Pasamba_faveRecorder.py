import openpyxl as op
from datetime import datetime

workbook = op.Workbook()
sheet = workbook.active

sheet['A1'] = "ID"
sheet['B1'] = "First Name"
sheet['C1'] = "Last Name"
sheet['D1'] = "Birth Year"
sheet['E1'] = "Age"

current_year = datetime.now().year
print("Person 1")
fname1 = input("Enter first name: ")
lname1 = input("Enter last name: ")
birth1 = int(input("Enter birth year: "))
age1 = current_year - birth1

print("\nPerson 2")
fname2 = input("Enter first name: ")
lname2 = input("Enter last name: ")
birth2 = int(input("Enter birth year: "))
age2 = current_year - birth2

print("\nPerson 3")
fname3 = input("Enter first name: ")
lname3 = input("Enter last name: ")
birth3 = int(input("Enter birth year: "))
age3 = current_year - birth3

sheet['A2'] = 1
sheet['B2'] = fname1
sheet['C2'] = lname1
sheet['D2'] = birth1
sheet['E2'] = age1

sheet['A3'] = 2
sheet['B3'] = fname2
sheet['C3'] = lname2
sheet['D3'] = birth2
sheet['E3'] = age2

sheet['A4'] = 3
sheet['B4'] = fname3
sheet['C4'] = lname3
sheet['D4'] = birth3
sheet['E4'] = age3

workbook.save("favorite_people.xlsx")
print("\nFavorite people saved successfully!")
wbk = op.load_workbook("favorite_people.xlsx")
sheet = wbk.active
print("\n==== FAVORITE PEOPLE LIST ====\n")
for rows in sheet.iter_rows(values_only=True):
    print(rows)