import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op

# --- Functions ---

def display():
    try:
        workbook = op.load_workbook("excelDB.xlsx")
        sheet = workbook.active
        # Clear table
        for row in table.get_children():
            table.delete(row)
        # Populate table
        for row in sheet.iter_rows(min_row=2, values_only=True):
            table.insert("", tk.END, values=row)
    except FileNotFoundError:
        # Create file if it doesn't exist
        wb = op.Workbook()
        ws = wb.active
        ws.append(["ID", "Last Name", "First Name", "Middle Name", "Birth Year", "Age"])
        wb.save("excelDB.xlsx")

def input_validation():
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    if not first or not middle or not last or not birth:
        messagebox.showerror("Error", "All fields required!")
        return False
    if not birth.isdigit():
        messagebox.showerror("Error", "Birthyear must be a number")
        return False
    return True

def saving():
    if not input_validation():
        return

    fname = fname_entry.get()
    mname = mname_entry.get()
    lname = lname_entry.get()
    by = int(birth_entry.get())
    age = 2026 - by  # Based on your system clock setting in the prompt

    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active
    
    # Simple ID generation based on row count
    new_id = sheet.max_row 
    
    sheet.append([new_id, lname, fname, mname, by, age])
    workbook.save("excelDB.xlsx")
    
    messagebox.showinfo("Success", "Record added successfully!")
    clear_entries()
    display()

def auto_populate(event):
    selected = table.focus()
    values = table.item(selected, "values")
    
    if values:
        clear_entries()
        # Mapping based on your table: ID(0), Last(1), First(2), Middle(3), Birth(4), Age(5)
        lname_entry.insert(0, values[1])
        fname_entry.insert(0, values[2])
        mname_entry.insert(0, values[3])
        birth_entry.insert(0, values[4])

def update():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Select a record first.")
        return
    
    if not input_validation():
        return

    values = table.item(selected, "values")
    record_id = values[0]
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = int(birth_entry.get())
    age = 2026 - birth

    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == str(record_id):
            row[1].value = last
            row[2].value = first
            row[3].value = middle
            row[4].value = birth
            row[5].value = age
            break

    workbook.save("excelDB.xlsx")
    messagebox.showinfo("Success", "Record updated successfully!")
    display()

def delete_record():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Select a record to delete.")
        return
    
    if messagebox.askyesno("Confirm", "Are you sure you want to delete this record?"):
        values = table.item(selected, "values")
        record_id = values[0]
        
        workbook = op.load_workbook("excelDB.xlsx")
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == str(record_id):
                sheet.delete_rows(row[0].row, 1)
                break
        
        workbook.save("excelDB.xlsx")
        display()
        clear_entries()

def clear_entries():
    fname_entry.delete(0, tk.END)
    mname_entry.delete(0, tk.END)
    lname_entry.delete(0, tk.END)
    birth_entry.delete(0, tk.END)

# --- UI Setup ---

window = tk.Tk()
window.title("Age Calculator")
window.geometry("800x600")
window.configure(bg="lightgreen")

# Title
title_label = tk.Label(window, text="Profile Builder", font=("Poppins", 16, "bold"), bg="lightgreen")
title_label.grid(row=0, column=0, columnspan=4, pady=10)

# Entry Frame for better organization
genframe = tk.Frame(window, bg="lightgreen")
genframe.grid(row=1, column=0, columnspan=4, padx=20)

# Input Fields
fname_entry = tk.Entry(genframe, font=("Poppins", 12))
fname_entry.grid(row=0, column=0, padx=5)
tk.Label(genframe, text="First Name", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=1, column=0)

mname_entry = tk.Entry(genframe, font=("Poppins", 12))
mname_entry.grid(row=0, column=1, padx=5)
tk.Label(genframe, text="Middle Name", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=1, column=1)

lname_entry = tk.Entry(genframe, font=("Poppins", 12))
lname_entry.grid(row=0, column=2, padx=5)
tk.Label(genframe, text="Last Name", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=1, column=2)

birth_entry = tk.Entry(window, font=("Poppins", 12))
birth_entry.grid(row=2, column=0, pady=(10, 0))
tk.Label(window, text="Birth Year", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=3, column=0)

# Buttons
btn_frame = tk.Frame(window, bg="lightgreen")
btn_frame.grid(row=4, column=0, columnspan=4, pady=20)

tk.Button(btn_frame, text="Submit", command=saving, bg="lightpink", font=("Poppins", 12, "bold")).pack(side=tk.LEFT, padx=10)
tk.Button(btn_frame, text="Update", command=update).pack(side=tk.LEFT, padx=10)
tk.Button(btn_frame, text="Delete", command=delete_record, bg="red", fg="white").pack(side=tk.LEFT, padx=10)

# Treeview Table
cols = ("ID", "Last Name", "First Name", "Middle Name", "Birth Year", "Age")
table = ttk.Treeview(window, columns=cols, show="headings")

for col in cols:
    table.heading(col, text=col)
    table.column(col, width=100)

table.grid(row=5, column=0, columnspan=4, padx=10, sticky='nsew')
table.bind("<<TreeviewSelect>>", auto_populate)

# Load data on start
display()

window.mainloop()
